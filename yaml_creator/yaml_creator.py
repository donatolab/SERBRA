import yaml
import os
import sys
import copy
import shutil
from datetime import datetime
from openpyxl import load_workbook, Workbook
import re
import numpy as np
import h5py

module_path = os.path.abspath(os.path.join("../"))
sys.path.append(module_path)
from Helper import *

manually_eddited_animals_yaml_fname = "animal_summary.yaml"


def row_to_list(sheet, row):
    result = []
    # Iterate over cells in the specified row
    for num, cell in enumerate(sheet[row][1:10000]):
        value = cell.value
        if cell.value != None:
            result.append(value)
        else:
            break
    return result


def num_to_date(date_string):
    if type(date_string) != str:
        date_string = str(date_string)
    date = datetime.strptime(date_string, "%Y%m%d")
    return date

def define_metadata_columns(sheet):
    metadata_columns = {}
    for column in range(1, sheet.max_column):
        cell = sheet.cell(row=1, column=column)
        if cell.value is not None:
            metadata_columns[cell.value] = column
    return metadata_columns

def create_stimulus_dict(sheet, metadata_columns, row, definition=None):
    # Stimulus Metadata
    stim_type = sheet.cell(row=row, column=metadata_columns["treadmill"]).value
    if definition is None:
        definition = {
            stim_type: {
                "sequence": None,
                "dimensions": None,
                "by": None,
            }
        }
    sequence = definition[stim_type]["sequence"] if "sequence" in definition[stim_type].keys() else None
    dimensions = definition[stim_type]["dimensions"] if "dimensions" in definition[stim_type].keys() else None
    by = definition[stim_type]["by"] if "by" in definition[stim_type].keys() else None
    metadata = create_dict(type=stim_type, sequence=sequence, dimensions=dimensions, by=by)
    return metadata


def create_behavior_dict(sheet, metadata_columns, row, stimulus_definition=None):
    # Behavior Metadata
    cam_data = sheet.cell(row=row, column=metadata_columns["cam"]).value
    cam_data = True if cam_data == "yes" else False
    movement_data = sheet.cell(row=row, column=metadata_columns["behaviour"]).value
    movement_data = True if movement_data == "yes" else False
    stimulus = create_stimulus_dict(sheet, metadata_columns, row, stimulus_definition)
    behavior_metadata = create_dict(
        cam_data=cam_data,
        movement_data=movement_data,
        stimulus=stimulus
    )
    return behavior_metadata

def create_neural_dict(sheet, metadata_columns, row):
    ## Neural Metadata
    method = "2P"
    setup = sheet.cell(row=row, column=metadata_columns["setup"]).value
    wavelength = sheet.cell(row=row, column=metadata_columns["[nm]"]).value
    laser_power = sheet.cell(row=row, column=metadata_columns["laser / LED power"]).value
    pockel_cell_bias = sheet.cell(row=row, column=metadata_columns["PC bias"]).value
    n_channel = sheet.cell(row=row, column=metadata_columns["n Ch."]).value
    functional_channel = sheet.cell(row=row, column=metadata_columns["fct. channel"]).value
    ug_gain = sheet.cell(row=row, column=metadata_columns["UG gain"]).value
    ur_gain = sheet.cell(row=row, column=metadata_columns["UR gain"]).value
    lens = sheet.cell(row=row, column=metadata_columns["lens"]).value if "lens" in metadata_columns.keys() else None
    pixels = sheet.cell(row=row, column=metadata_columns["pixels"]).value if "pixels" in metadata_columns.keys() else None
    n_planes = sheet.cell(row=row, column=metadata_columns["n planes"]).value if "n planes" in metadata_columns.keys() else None

    convert_to_int = [n_channel, functional_channel, n_planes]
    for i, value in enumerate(convert_to_int):
        if value:
            if value == "n/a" or value == "" or value == "?":
                convert_to_int[i] = None
            else:
                convert_to_int[i] = int(value)
    n_channel, functional_channel, n_planes = convert_to_int

    neural_metadata = create_dict(
        method=method,
        setup=setup,
        wavelength=wavelength,
        laser_power=laser_power,
        pockel_cell_bias=pockel_cell_bias,
        n_channel=n_channel,
        functional_channel=functional_channel,
        ug_gain=ug_gain,
        ur_gain=ur_gain,
        lens=lens,
        pixels=pixels,
        n_planes=n_planes,
    )
    return neural_metadata

def create_task_dict(sheet, metadata_columns, row, stimulus_definition=None):
    # Task Metadata
    duration = sheet.cell(row=row, column=metadata_columns["duration [min]"]).value
    duration = None if duration == "n/a" or duration == "" or duration == "?" else duration
    expt_pipeline = sheet.cell(row=row, column=metadata_columns["paradigm"]).value
    comment = sheet.cell(row=row, column=metadata_columns["comment"]).value
    task_metadata = create_dict(expt_pipeline=expt_pipeline, comment=comment, duration=duration)
    neural_metadata = create_neural_dict(sheet, metadata_columns, row)
    behavior_metadata = create_behavior_dict(sheet, metadata_columns, row, stimulus_definition)
    task_metadata["neural_metadata"] = neural_metadata
    task_metadata["behavior_metadata"] = behavior_metadata
    return task_metadata

def create_session_dict(sheet, metadata_columns, row, stimulus_definition=None):
    # Session Metadata
    date = sheet.cell(row=row, column=metadata_columns["date"]).value
    date = "20" + str(int(date)) if date else None
    weight = sheet.cell(row=row, column=metadata_columns["weight [g]"]).value
    session_metadata = create_dict(date=date, weight=weight)
    task_metadata = create_task_dict(sheet, metadata_columns, row, stimulus_definition)
    task = sheet.cell(row=row, column=metadata_columns["session"]).value
    session_metadata["tasks_infos"] = {task: task_metadata}
    return session_metadata

def create_animal_dict(sheet, metadata_columns, row, stimulus_definition=None):
    animal_id = sheet.cell(row=row, column=metadata_columns["mouse ID"]).value
    if not animal_id:
        return None
    else:
        animal_id = (
            "DON-00" + animal_id[3:]
            if len(animal_id) == 7
            else "DON-0" + animal_id[3:]
        )
    sex = sheet.cell(row=row, column=metadata_columns["sex"]).value
    sex = "male" if sex == "m" else "female" if sex else None

    dob = sheet.cell(row=row, column=metadata_columns["DOB"]).value
    dob = "20" + str(int(dob)) if dob else None
    # dob_date = num_to_date(dob)
    # cohort_year = dob_date.year

    injected = sheet.cell(row=row, column=metadata_columns["injected"]).value
    injected = "20" + str(int(injected)) if injected else None
    implanted = sheet.cell(row=row, column=metadata_columns["implanted"]).value
    implanted = "20" + str(int(implanted)) if implanted else None
    
    animal_metadata = create_dict(animal_id=animal_id, sex=sex, dob=dob, injected=injected, implanted=implanted)
    session_metadata = create_session_dict(sheet, metadata_columns, row, stimulus_definition)
    animal_metadata["sessions"] = {session_metadata["date"]: session_metadata}
    return animal_metadata

def get_animal_dict_from_spreadsheet(fname, sheet_title=None, stimulus_definition=None, remove_none=True):
    """
    read animal_id, dob, sex from spreadsheet
    """
    org_exp_workbook = load_workbook(filename=fname)
    sheet_title = sheet_title if sheet_title else org_exp_workbook.sheetnames[0]
    if sheet_title not in org_exp_workbook.sheetnames:
        raise ValueError(f"sheet_title {sheet_title} not in sheetnames in {fname}")
    sheet = org_exp_workbook[sheet_title]
    metadata_columns = define_metadata_columns(sheet)
    animals = {}
    for row in range(2, sheet.max_row):
        row_animal_metadata = create_animal_dict(sheet, metadata_columns, row, stimulus_definition)
        if row_animal_metadata is None:
            continue

        #session_metadata = create_dict(sheet, metadata_columns, row)
        #task_metadata = create_task_dict(sheet, metadata_columns, row)
        #neural_metadata = create_neural_dict(sheet, metadata_columns, row)
        #behavior_metadata = create_behavior_dict(sheet, metadata_columns, row, stimulus_definition)

        animal_id = row_animal_metadata["animal_id"]
        session_date = row_animal_metadata["sessions"][list(row_animal_metadata["sessions"].keys())[0]]["date"]
        
        if len(animals.keys()) > 0:
            if animal_id in animals:
                animal = animals[animal_id]
                animal_exists = True
                if session_date in animal["sessions"]:
                    session = animal["sessions"][session_date]
                    session["tasks_infos"].update(row_animal_metadata["sessions"][session_date]["tasks_infos"])
                else:
                    animal["sessions"].update(row_animal_metadata["sessions"])
            else:
                animals[animal_id] = row_animal_metadata
        else:
            animals[animal_id] = row_animal_metadata

    if remove_none:
        animals = remove_none_from_dict(animals, recursive=True)
    return animals


def create_dict(**kwargs):
    session_dict = kwargs
    for key, var in session_dict.items():
        session_dict[key] = None if var == "n/a" or var == "" or var == "?" else var
    # WARNING dob_date.year could be wrong for other
    return session_dict

def search_update_dict(dictionary, update_dict):
    for key, value in update_dict.items():
        if key in dictionary.keys():
            dictionary_value = dictionary[key]
            if isinstance(dictionary_value, dict):
                dictionary_value.update(value)
            
        else:
            for dict_key, dict_value in dictionary.items():
                if isinstance(dict_value, dict):
                    search_update_dict(dict_value, update_dict)
    return dictionary

def remove_none_from_dict(dictionary, recursive=False):
    """
    removes all None values from a dictionary
    parameters:
        dictionary (dict): dictionary to remove None values from
        recursive (bool): if True, also removes None values from nested dictionaries
    """
    if not recursive:
        new_dict = {key: value for key, value in dictionary.items() if value is not None}
    else:
        new_dict = {}
        for key, value in dictionary.items():
            if value is not None:
                if isinstance(value, dict):
                    new_dict[key] = remove_none_from_dict(value, recursive)
                else:
                    new_dict[key] = value
    return new_dict


def return_loaded_yaml_if_newer(used_path, may_newer_info_path):
    yaml_dict = None
    root_yaml_modification_date = (
        os.path.getmtime(used_path) if os.path.exists(used_path) else 0
    )
    if os.path.exists(may_newer_info_path):
        yaml_modification_date = os.path.getmtime(may_newer_info_path)
        if yaml_modification_date > root_yaml_modification_date:
            with open(may_newer_info_path, "r") as yaml_file:
                yaml_dict = yaml.safe_load(yaml_file)
    return yaml_dict


def get_animals_from_yaml(directory):
    root_dir = directory if directory else ""
    root_yaml_path = os.path.join(root_dir, manually_eddited_animals_yaml_fname)
    if os.path.exists(root_yaml_path):

        with open(root_yaml_path, "r") as yaml_file:
            animals = yaml.safe_load(yaml_file)
    else:
        animals = {}

    """
    for animal_id in get_directories(root_dir, regex_search="DON-"):
        animal_path = os.path.join(root_dir, animal_id)

        # update animals if a file has newer information changed
        animal_yaml_path = os.path.join(animal_path, animal_id+".yaml")
        animal = return_loaded_yaml_if_newer(root_yaml_path, animal_yaml_path)
        animal
        animals[animal_id] = animal if animal else animals[animal_id]
    
    to_delete_animals = []
    to_delete_keys = ["pdays", "cohort_year", "functional_channels", "sex", "session_dates", "session_names", "dob"]
    for animal_id, animal in animals.items():
        if 'UseMUnits' not in animal.keys():
            to_delete_animals.append(animal_id)
            continue
        for to_delete_key in to_delete_keys:
            if to_delete_key in animal.keys():# and 'UseMUnits' in animal.keys():
                del animal[to_delete_key]
    for to_delete_animal in to_delete_animals:
        del animals[to_delete_animal]"""
    return animals


def combine_spreadsheet_and_old_animal_summary_yaml(animals_spreadsheet, animals_yaml):
    import copy

    for yanimal_id, yanimal in animals_yaml.items():
        for date, ysession in yanimal["sessions"].items():

            if date not in animals_spreadsheet[yanimal_id]["sessions"].keys():

                animals_spreadsheet[yanimal_id]["sessions"][date] = create_dict(
                    date=date, method="2P", setup="femtonics"
                )

            if "UseMUnits" in ysession.keys():
                usemunits = ysession["UseMUnits"]
                if not usemunits:
                    continue
                else:
                    animals_spreadsheet[yanimal_id]["sessions"][date][
                        "UseMUnits"
                    ] = usemunits
    return animals_spreadsheet


def add_session_animal_folders(animals, animals_spreadsheet, directory=None):
    root_dir = directory if directory else ""
    for animal_id in get_directories(root_dir, regex_search="DON-"):
        if animal_id not in animals:
            animals[animal_id] = animals_spreadsheet[animal_id]
            print(f"added animal from spreadsheet: {animal_id}")
        animal_path = os.path.join(root_dir, animal_id)

        functional_channel_list = []
        for session_id in get_directories(animal_path):
            session_path = os.path.join(animal_path, session_id, "002P-F")

            mesc_fnames = get_files(session_path, ending=".mesc")
            mesc_munit_pairs = (
                animals[animal_id]["UseMUnits"]
                if "UseMUnits" in animals[animal_id].keys()
                else []
            )
            for fname in mesc_fnames:
                splitted_fname = fname.split("_")
                if animal_id != splitted_fname[0]:
                    continue

                # add session data based on file
                session_date = splitted_fname[1]
                if (
                    session_id not in animals[animal_id]["session_names"]
                    and session_date not in animals[animal_id]["session_dates"]
                ):
                    animals[animal_id]["session_names"].append(session_id)
                    animals[animal_id]["session_dates"].append(session_date)
                    dob_date = num_to_date(animals[animal_id]["dob"])
                    session_date = num_to_date(session_date)
                    pday = (session_date - dob_date).days
                    animals[animal_id]["pdays"].append(pday)

                # get available MUnits
                last_fname_part = splitted_fname[-1].split(".")[0]
                session_parts = [
                    int(part_number[-1]) - 1
                    for part_number in re.findall("S[0-9]", last_fname_part)
                ]
                fpath = os.path.join(session_path, fname)
                munits_list, number_channels = get_recording_munits(
                    fpath, session_parts
                )

                # Get MUnit number list of first Mescfile session MSession_0
                if len(munits_list) <= len(session_parts):
                    usefull_munits = munits_list
                    file_naming = session_parts[: len(usefull_munits)]
                else:
                    add_mesc_munit_pair = True
                    if mesc_munit_pairs:
                        if len(mesc_munit_pairs) > 0:
                            for mesc_munit_pair in mesc_munit_pairs:
                                if fname in mesc_munit_pair:
                                    add_mesc_munit_pair = False
                        if add_mesc_munit_pair:
                            mesc_munit_pair = [fname, munits_list]
                            mesc_munit_pairs.append(mesc_munit_pair)

                functional_channel = (
                    2
                    if 20210821 < int(session_date)
                    and int(session_date) < 20220422
                    and number_channels > 1
                    else 1
                )
                functional_channel_list.append(functional_channel)

            # TODO: integrate functional channels into individual yaml files
            animals[animal_id]["functional_channels"] = functional_channel_list

            if mesc_munit_pairs:
                if len(mesc_munit_pairs) > 0:
                    animals[animal_id]["UseMUnits"] = mesc_munit_pairs
    return animals


def get_recording_munits(
    mesc_fpath, session_parts, fps=30, at_least_minutes_of_recording=5
):
    # Get MUnit number list of first Mescfile session MSession_0
    with h5py.File(mesc_fpath, "r") as file:
        munits = file[list(file.keys())[0]]
        recording_munits = []
        for name, unit in munits.items():
            # if recording has at least x minutes
            if unit["Channel_0"].shape[0] > fps * 60 * at_least_minutes_of_recording:
                unit_number = name.split("_")[-1]
                recording_munits.append(int(unit_number))
                # get number of imaging channels
                number_channels = 0
                for key in unit.keys():
                    if "Channel" in key:
                        number_channels += 1
    return recording_munits, number_channels


def move_mesc_to_session_folder(directory=None):
    directory = None if directory == "" else directory
    directory = Path(directory)
    for fname in get_files(directory, ending=".mesc"):
        splitted_fname = fname.split("_")
        if splitted_fname[0][:3] != "DON":  # not animal
            continue
        animal_id = splitted_fname[0]
        session_id = splitted_fname[1].split(".")[0]
        session_path = directory.joinpath(animal_id, session_id)
        session_path.mkdir(parents=True, exist_ok=True)
        # move_file
        fpath = os.path.join(directory, fname)
        shutil.move(fpath, session_path)

def create_folders_for_animals(animals, directory=None, save_yamls=True):
    directory = None if directory == "" else directory
    directory = Path(directory)
    for animal_id, animal_metadata in animals.items():
        animal_path = directory.joinpath(animal_id)
        for session_date, session_metadata in animal_metadata["sessions"].items():
            session_path = animal_path.joinpath(session_date)
            if save_yamls:
                fpath = os.path.join(session_path, f"{session_date}.yaml")
                with open(fpath, "w") as file:
                    yaml.dump(session_metadata, file)
        if save_yamls:
            fpath = os.path.join(animal_path, f"{animal_id}.yaml")
            only_animal_metadata = copy.deepcopy(animal_metadata)
            only_animal_metadata.pop("sessions")
            with open(fpath, "w") as file:
                yaml.dump(only_animal_metadata, file)

def update_excel_by_yaml(excel_animals, yaml_animals):
    pass
    # FIXME: continue


def main(directory=None):
    root_dir = directory if directory else ""
    fname = os.path.join("Intrinsic_CA3_database-September_7,_10_08_AM.xlsx")
    fpath = os.path.join(root_dir, fname)
    # load spreadsheet information
    animals_spreadsheet = get_animal_dict_from_spreadsheet(fpath)
    # move mesc in root directory to correct folder location
    move_mesc_to_session_folder(directory=root_dir)
    # load animal yaml files
    animals_yaml = get_animals_from_yaml(root_dir)
    animals = combine_spreadsheet_and_old_animal_summary_yaml(
        animals_spreadsheet, animals_yaml
    )
    # get animals based on folder structure
    # TODO: old stuff may create new version
    # print(ooold)
    # animals = add_session_animal_folders(
    #    animals_yaml, animals_spreadsheet, directory=root_dir
    # )
    add_yaml_to_folders(animals, directory=root_dir)
    # with open(os.path.join(root_dir, manually_eddited_animals_yaml_fname), "w") as file:
    #    yaml.dump(animals, file)
    # save yaml files in folders and root directory
    with open(os.path.join(root_dir, "animals.yaml"), "w") as file:
        yaml.dump(animals, file)


if __name__ == "__main__":
    root_dir = "/scicore/projects/donafl00-calcium/Users/Sergej/Steffen_Experiments"
    main()
